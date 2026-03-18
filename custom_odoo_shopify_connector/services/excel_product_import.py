"""
Import products from an Excel (.xlsx) file into Odoo.

Expected columns (first row = header, case-insensitive):
- Name (required)
- Description
- SKU / Internal Reference
- List Price / Sales Price / Price
- Cost
- Category
- Barcode
"""
import base64
import io
import logging

from odoo import _

_logger = logging.getLogger(__name__)

# Column header variants we accept (first match wins)
NAME_KEYS = ("name", "product name", "product")
DESC_KEYS = ("description", "desc")
SKU_KEYS = ("sku", "internal reference", "default code", "reference")
PRICE_KEYS = ("list price", "sales price", "price", "list_price", "sales_price")
COST_KEYS = ("cost", "standard price", "standard_price")
CATEGORY_KEYS = ("category", "categ", "product category")
BARCODE_KEYS = ("barcode",)


def _column_index(header_row, keys_tuple):
    """Return 0-based column index for a header that matches any of keys_tuple, or None."""
    for idx, cell in enumerate(header_row):
        val = (cell.value or "").strip().lower() if cell else ""
        for k in keys_tuple:
            if k in val or val in k:
                return idx
    return None


def import_products_from_excel(env, file_content_base64, filename=""):
    """
    Parse Excel file and create/update product.template and product.product.
    file_content_base64: base64-encoded file content.
    Returns (created_count, updated_count, errors).
    """
    try:
        import openpyxl
    except ImportError:
        raise ValueError(
            _("The openpyxl library is required for Excel import. Install it with: pip install openpyxl")
        )

    if not file_content_base64:
        raise ValueError(_("Please upload an Excel file."))

    raw = base64.b64decode(file_content_base64)
    if not raw:
        raise ValueError(_("The uploaded file is empty."))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        _logger.exception("Excel load failed")
        raise ValueError(_("Invalid or corrupted Excel file: %s") % e)

    ws = wb.active
    if not ws:
        raise ValueError(_("The Excel file has no active sheet."))

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row or 0))
    if len(rows) < 2:
        raise ValueError(_("The Excel file must have a header row and at least one data row."))

    header = rows[0]
    name_col = _column_index(header, NAME_KEYS)
    if name_col is None:
        raise ValueError(
            _("The first row must contain a 'Name' column (product name).")
        )

    desc_col = _column_index(header, DESC_KEYS)
    sku_col = _column_index(header, SKU_KEYS)
    price_col = _column_index(header, PRICE_KEYS)
    cost_col = _column_index(header, COST_KEYS)
    category_col = _column_index(header, CATEGORY_KEYS)
    barcode_col = _column_index(header, BARCODE_KEYS)

    ProductTemplate = env["product.template"]
    ProductCategory = env["product.category"]
    created = 0
    updated = 0
    errors = []

    for row_index, row in enumerate(rows[1:], start=2):
        try:
            name_val = row[name_col].value if name_col < len(row) else None
            name = (name_val or "").strip() if name_val is not None else ""
            if not name:
                continue

            def _cell(col):
                if col is None or col >= len(row):
                    return None
                v = row[col].value
                if v is None:
                    return None
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    return v
                return (v or "").strip() or None

            description = _cell(desc_col)
            sku = _cell(sku_col)
            list_price = _cell(price_col)
            cost = _cell(cost_col)
            category_name = _cell(category_col)
            barcode = _cell(barcode_col)

            if list_price is not None:
                try:
                    list_price = float(list_price)
                except (TypeError, ValueError):
                    list_price = 0.0
            else:
                list_price = 0.0

            if cost is not None:
                try:
                    cost = float(cost)
                except (TypeError, ValueError):
                    cost = 0.0
            else:
                cost = 0.0

            categ_id = False
            if category_name:
                cat = ProductCategory.search([("name", "=", category_name)], limit=1)
                if not cat:
                    cat = ProductCategory.create({"name": category_name})
                categ_id = cat.id

            # Find existing by SKU or by name
            variant = env["product.product"].browse()
            if sku:
                variant = env["product.product"].search([("default_code", "=", sku)], limit=1)
            template = variant.product_tmpl_id if variant else ProductTemplate.search([("name", "=", name)], limit=1)

            vals = {
                "name": name,
                "type": "consu",
                "list_price": list_price,
                "categ_id": categ_id or template.categ_id.id if template else False,
            }
            if description is not None:
                vals["description"] = description
            if cost is not None and cost != 0:
                vals["standard_price"] = cost

            if template:
                template.write(vals)
                updated += 1
                product = template.product_variant_id
            else:
                template = ProductTemplate.create(vals)
                product = template.product_variant_ids[0]
                created += 1

            # Update variant-level fields
            variant_vals = {}
            if sku is not None:
                variant_vals["default_code"] = sku or False
            if barcode is not None:
                variant_vals["barcode"] = barcode or False
            if variant_vals:
                product.write(variant_vals)

        except Exception as e:
            errors.append(_("Row %s: %s") % (row_index, e))
            _logger.warning("Excel row %s error: %s", row_index, e)

    wb.close()
    return created, updated, errors
